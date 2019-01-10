from urllib import request
from bs4 import BeautifulSoup
from openpyxl import Workbook
from articleTest import articleParser
from makeFolder import mkFolder
from articleTest import timeParser
from articleTest import typeParser
from articleTest import uidParser
import re

# times_re = re.compile('authorposton\d{7,9}')
# types_re = re.compile('.*&ac=usergroup&.*')
user_re = re.compile('space-uid-.*')
#Request the url and return the url
def requestUrl(index):
    response = request.urlopen(index)
    html = response.read()
    html = html.decode("utf-8")

#    print(html)
    return html

#parse the blog
def postParser(index, page_no):
#create the folder and the sheet
    class_name = "其他"
    file_name = class_name + str(page_no)
    dir_path = r"D:\MIUI" + "\\" + class_name
    mkFolder(dir_path)
    filepath = dir_path + "\\"+ file_name + ".xlsx"
    wb = Workbook()
    sheet = wb.create_sheet(class_name, index = 0)
    sheet.title = file_name

    #re.compile objects

    soup = BeautifulSoup(requestUrl(index),'html.parser');
    all_posts = soup.find_all('a', attrs={'class':'s xst'})
    all_users = soup.find_all('a', attrs={'href':user_re,'c':'1'})

    for post_index,post_item in enumerate(all_posts):
        article_result = list()
        uid_result =list()
        type_result = list()

        post_item_content = post_item.get_text()
        post_item_href = "http://www.miui.com/"+ post_item.get('href')
        article_result = articleParser(post_item_href)

        uid_result = uidParser(post_item_href)
        type_result = typeParser(post_item_href)

        post_time = timeParser(post_item_href)

        sheet['A' + str(post_index + 2)] = post_item_content
        sheet['C'+ str(post_index + 2)] = post_time
        for result_index in range(len(article_result)):
            sheet.cell(row = post_index + 2, column = 4 + result_index).value = \
            uid_result[result_index]+'\\'+\
            type_result[result_index]+'\\'+\
            article_result[result_index]

        # print(post_item_content)

    for user_index,user_item in enumerate(all_users):
        user_item_name = user_item.get_text()
        # user_item_modified_name = re.sub("\n","",user_item_name)
        sheet['B' + str(user_index + 2)] = user_item_name
        # if user_item_modified_name is not None:
        #     print(user_item_modified_name)

    wb.save(filepath)
#
# if __name__ == "__main__":
#     extract_index = "http://www.miui.com/type-38-7467.html"
#     postParser(extract_index,1);
